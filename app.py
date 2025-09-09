"""
Streamlit App: Research Platform Picker (Beginner-Friendly)

What this app does
------------------
- Loads an Excel file named "Research Questionnaire Info.xlsx" where:
    - Row 4 contains the column headers (so data starts on row 5).
    - Columns A through G are the input criteria.
    - Column H is the output/label for that row (e.g., recommended platform).
    - Cells in columns A–G are color-coded:
        - GREEN means the comparison operator is "<=" (user_input <= cell_value)
        - RED   means the comparison operator is ">"  (user_input > cell_value)
        - If a cell has no fill color, we default to "<=" (beginner-friendly fallback).
- Lets a user enter their own inputs for each criterion.
- **New:** Three criteria are treated as categorical with dropdowns instead of numbers:
    - **OS** → options: `Windows`, `Linux`, `any`
    - **Data Sensitivity** → options: `Open`, `ITAR`, `PHI`
    - **Software** → options: `ANSYS-CFD`, `other`, `any`, `MATLAB`, `COMSOL`, `open-source`
- Scores each row equally (all criteria have the same weight) based on how many criteria match.
- Ranks rows by match score and suggests the top rows.

How to run this app
-------------------
1) Install requirements (in a virtual environment is recommended):

    pip install streamlit pandas openpyxl numpy

2) Put your Excel file in the same folder as this script with the exact name:

    Research Questionnaire Info.xlsx

3) Run the Streamlit app:

    streamlit run streamlit_research_platform_picker.py

4) A browser tab will open. Enter your inputs in the sidebar, then review the ranked results.

Tip: You can also upload a different Excel file in the sidebar if needed.
"""

import io
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# -----------------------------
# Constants for categorical criteria
# -----------------------------
CATEGORICAL_SPECS = {
    # key should match the Excel column header exactly (case-insensitive compare is used)
    "os": {
        "options": ["Windows", "Linux", "any"],
        "has_any": True,
    },
    "data sensitivity": {
        "options": ["Open", "ITAR", "PHI"],
        "has_any": False,
    },
    "software": {
        "options": ["ANSYS-CFD", "other", "any", "MATLAB", "COMSOL", "open-source"],
        "has_any": True,
    },
}

# -----------------------------
# Utility: detect cell fill color
# -----------------------------
# Excel colors can vary, but common RGB values:
#   - Red often appears like 'FFFF0000' or 'FF0000'
#   - Green often appears like 'FF00B050' or 'FF92D050' or 'FF00FF00'
# We'll use a simple heuristic to classify fills as red/green. If neither, return None.

def _argb_to_str(color) -> str:
    """Convert an openpyxl color to an uppercase ARGB hex string if possible."""
    if color is None:
        return ""
    try:
        if color.type == "rgb" and color.rgb:
            return str(color.rgb).upper()
    except Exception:
        pass
    return ""


def _classify_fill(cell) -> str:
    """Return 'green', 'red', or '' (unknown) for a cell's fill color."""
    fill = getattr(cell, "fill", None)
    if not fill or not getattr(fill, "start_color", None):
        return ""
    argb = _argb_to_str(fill.start_color)
    if not argb:
        return ""

    argb = argb.upper()
    # Red-like
    if any(s in argb for s in ["FF0000", "FF5B5B", "FFC000"]):  # include strong reds / orange-ish
        return "red"
    # Green-like
    if any(s in argb for s in ["00B050", "92D050", "00FF00", "66FF66"]):
        return "green"
    return ""


# -----------------------------
# Loaders
# -----------------------------

def load_excel_with_colors(file_bytes: bytes) -> Tuple[pd.DataFrame, List[List[str]]]:
    """Load the Excel file (headers at row 4) as a DataFrame and also return a matrix
    of color tags ("green"/"red"/"") for columns A–G.

    Returns:
        df: pandas DataFrame with header=3 (0-based), so headers from row 4 in Excel
        color_matrix: list of lists, shape (n_rows, 7), colors for A..G for each row
    """
    df = pd.read_excel(io.BytesIO(file_bytes), header=3, engine="openpyxl")

    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active  # first sheet by default

    start_row = 5  # data begins on row 5 in Excel (1-based)
    color_matrix: List[List[str]] = []

    for i in range(len(df)):
        excel_row = start_row + i
        row_colors: List[str] = []
        for col_idx in range(1, 8):  # columns A..G
            cell = ws.cell(row=excel_row, column=col_idx)
            c = _classify_fill(cell)
            row_colors.append(c)
        color_matrix.append(row_colors)

    return df, color_matrix


# -----------------------------
# Matching & Scoring
# -----------------------------

def build_comparison_ops(color_row: List[str]) -> List[str]:
    """Map color tags to comparison operators.
    For numeric columns:
      GREEN -> '<='
      RED   -> '>'
      ''    -> '<=' (fallback)
    For categorical columns we'll override to 'cat' later.
    """
    ops = []
    for c in color_row:
        if c == "green":
            ops.append("<=")
        elif c == "red":
            ops.append(">")
        else:
            ops.append("<=")  # default fallback
    return ops


def to_number_or_nan(x):
    """Try to coerce to float, otherwise return NaN."""
    try:
        if pd.isna(x):
            return np.nan
        return float(x)
    except Exception:
        return np.nan


def _tokenize(value: str) -> List[str]:
    """Tokenize a categorical cell value like 'Windows/Linux, any' into lowercase tokens.
    Recognizes separators: ',', '/', ';'.
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return []
    s = str(value).strip().lower()
    if not s:
        return []
    # split on common separators
    for sep in [",", "/", ";"]:
        s = s.replace(sep, "|")
    toks = [t.strip() for t in s.split("|") if t.strip()]
    return toks


def compare_categorical(user_choice: str, row_value: str, has_any: bool) -> bool:
    """Compare dropdown choice against the row's (possibly multi-valued) entry.
    Rules:
      - If user_choice == 'any' and the spec allows 'any' → always pass.
      - If row_value contains 'any' (case-insensitive) → pass.
      - Otherwise, pass if user_choice is contained in row_value tokens (case-insensitive).
    """
    uc = (user_choice or "").strip().lower()
    row_tokens = set(_tokenize(row_value))

    if has_any and uc == "any":
        return True
    if "any" in row_tokens:
        return True
    return uc in row_tokens


def evaluate_row(user_inputs: List, row_values: List, ops: List[str], col_names: List[str]) -> Tuple[int, int, List[bool]]:
    """Compare user inputs to a single row's criteria using the operators.

    Args:
        user_inputs: list of values for columns A..G (length 7); may contain floats or strings
        row_values:  list of values from the DataFrame for columns A..G; may contain floats or strings
        ops:         per-column operators: '<=', '>' for numeric, or 'cat' for categorical
        col_names:   list of column names for A..G

    Returns:
        (num_matches, total_considered, pass_list)
    """
    matches = 0
    considered = 0
    pass_list: List[bool] = []

    for val_user, val_row, op, name in zip(user_inputs, row_values, ops, col_names):
        # Categorical comparison
        if op == "cat":
            spec = CATEGORICAL_SPECS.get(name.lower(), {"has_any": False})
            ok = compare_categorical(str(val_user), str(val_row), has_any=spec.get("has_any", False))
            considered += 1  # always considered if we have a user choice
            if ok:
                matches += 1
            pass_list.append(ok)
            continue

        # Numeric comparison
        u = to_number_or_nan(val_user)
        r = to_number_or_nan(val_row)
        if np.isnan(u) or np.isnan(r):
            pass_list.append(False)
            continue
        considered += 1
        if op == "<=":
            ok = u <= r
        else:  # op == '>'
            ok = u > r
        if ok:
            matches += 1
        pass_list.append(ok)

    return matches, considered, pass_list


# -----------------------------
# Streamlit UI
# -----------------------------

def main():
    st.set_page_config(page_title="Research Platform Picker", page_icon="🔎", layout="wide")
    st.title("🔎 Research Platform Picker")

    # Sidebar configuration and file input
    st.sidebar.header("1) Data Source")
    st.sidebar.write(
        "By default, the app loads **Research Questionnaire Info.xlsx** from this folder."
        "You can also upload a file with the same structure."
    )

    uploaded = st.sidebar.file_uploader("Upload Excel (optional)", type=["xlsx"])  # optional

    # Try to load either the uploaded file or the default file path
    excel_bytes: bytes = b""
    load_error = None
    if uploaded is not None:
        try:
            excel_bytes = uploaded.read()
        except Exception as e:
            load_error = f"Error reading uploaded file: {e}"
    else:
        try:
            with open("Research Questionnaire Info.xlsx", "rb") as f:
                excel_bytes = f.read()
        except FileNotFoundError:
            load_error = (
                "Could not find 'Research Questionnaire Info.xlsx' in the current folder.
"
                "Please upload a file using the sidebar."
            )
        except Exception as e:
            load_error = f"Error opening default Excel file: {e}"

    if load_error:
        st.error(load_error)
        st.stop()

    # Load the Excel into DataFrame and color matrix
    try:
        df, color_matrix = load_excel_with_colors(excel_bytes)
    except Exception as e:
        st.error(f"Failed to parse Excel: {e}")
        st.stop()

    if df.shape[1] < 8:
        st.error("Expected at least 8 columns (A..H). Please check the Excel structure.")
        st.stop()

    # Keep only the first 8 columns in case there are extras (A..H)
    df = df.iloc[:, :8].copy()

    # Column A..G are criteria; H is output label (e.g., platform)
    criteria_cols = df.columns[:7]  # A..G
    output_col = df.columns[7]      # H

    # Determine which criteria are categorical by header name
    is_categorical = [col.lower() in CATEGORICAL_SPECS for col in criteria_cols]

    # Show a small preview to the user
    with st.expander("Preview data (first 10 rows)"):
        st.dataframe(df.head(10))
        st.caption(
            "Note: Headers are taken from row 4 of the Excel file. "
            "Columns A–G are treated as criteria; column H is the output label."
        )

    st.sidebar.header("2) Enter Your Inputs")
    st.sidebar.write(
        "Provide your values for each criterion below.
"
        "Cells colored **green** in Excel mean your input should be **<=** the value in that cell.
"
        "Cells colored **red** mean your input should be **>** the value in that cell.
"
        "For categorical fields (OS, Data Sensitivity, Software), select from the dropdown."
    )

    # Build inputs with sensible defaults
    user_inputs: List = []

    # Numeric defaults: use column medians
    numeric_defaults = df[criteria_cols].apply(pd.to_numeric, errors="coerce").median(numeric_only=True)

    for col in criteria_cols:
        col_key = col.lower()
        if col_key in CATEGORICAL_SPECS:
            spec = CATEGORICAL_SPECS[col_key]
            # sensible defaults: 'any' when available, else first option
            if spec.get("has_any", False):
                default_choice = "any"
            else:
                default_choice = spec["options"][0]
            val = st.sidebar.selectbox(
                f"{col}",
                options=spec["options"],
                index=spec["options"].index(default_choice) if default_choice in spec["options"] else 0,
                help="Select your preference for this criterion",
            )
            user_inputs.append(val)
        else:
            default_val = numeric_defaults.get(col, np.nan)
            default_val = float(default_val) if not np.isnan(default_val) else 0.0
            val = st.sidebar.number_input(
                f"{col}",
                value=float(default_val),
                help="Your desired value for this numeric criterion",
            )
            user_inputs.append(float(val))

    st.sidebar.header("3) Scoring Options")
    normalize = st.sidebar.checkbox(
        "Normalize scores by number of comparable criteria",
        value=True,
        help=(
            "If checked, the score is matches / comparable_criteria. "
            "If unchecked, the score is just the count of matches."
        ),
    )

    st.sidebar.info(
        "All criteria are equally weighted. The top result is the one with the highest score."
    )

    # Evaluate each row
    results: List[Dict] = []

    # Prepare row values: keep strings for categorical columns, numerics for numeric columns
    crit_values_all = []
    for i, col in enumerate(criteria_cols):
        if col.lower() in CATEGORICAL_SPECS:
            crit_values_all.append(df[col].astype(str).fillna(""))
        else:
            crit_values_all.append(pd.to_numeric(df[col], errors="coerce"))
    crit_values_df = pd.concat(crit_values_all, axis=1)
    crit_values_df.columns = list(criteria_cols)

    # Build base ops from colors, then override to 'cat' for categorical columns
    ops_matrix: List[List[str]] = []
    for i in range(len(df)):
        ops = build_comparison_ops(color_matrix[i])
        ops = ["cat" if criteria_cols[j].lower() in CATEGORICAL_SPECS else ops[j] for j in range(7)]
        ops_matrix.append(ops)

    for i in range(len(df)):
        row_vals = [crit_values_df.iloc[i][c] for c in criteria_cols]
        ops = ops_matrix[i]
        matches, considered, pass_list = evaluate_row(user_inputs, row_vals, ops, list(criteria_cols))
        score = (matches / considered) if normalize and considered > 0 else matches

        results.append(
            {
                "Row": i + 1,
                "Output": df.iloc[i][output_col],
                "Matches": matches,
                "Comparable Criteria": considered,
                "Score": score,
                "Pass List": pass_list,
                "Ops": ops,
                "Row Values": row_vals,
            }
        )

    if not results:
        st.warning("No rows to evaluate.")
        st.stop()

    res_df = pd.DataFrame(results)
    res_df_sorted = res_df.sort_values(by=["Score", "Matches"], ascending=False).reset_index(drop=True)

    st.subheader("Results")

    # Show top 5 summary table (excluding verbose columns)
    summary_cols = ["Output", "Matches", "Comparable Criteria", "Score", "Row"]
    st.dataframe(res_df_sorted[summary_cols].head(5))

    # Detailed explanation for the best match
    top = res_df_sorted.iloc[0]
    st.markdown("### 🏆 Top Match")
    st.write(f"**Output (Column H):** {top['Output']}")
    st.write(f"**Score:** {top['Score']:.3f} ({int(top['Matches'])} of {int(top['Comparable Criteria'])} criteria matched)")

    with st.expander("Why this row scored as it did (detailed breakdown)"):
        # Display 'cat' for categorical columns, else the numeric operator
        display_ops = [op if op != "cat" else "categorical match" for op in top["Ops"]]
        detail = pd.DataFrame(
            {
                "Criterion": list(criteria_cols),
                "Your Input": user_inputs,
                "Operator": display_ops,
                "Row Value": top["Row Values"],
                "Pass?": top["Pass List"],
            }
        )
        st.dataframe(detail)

    st.caption(
        "Notes: GREEN cells mean '<=', RED cells mean '>'. For categorical columns (OS, Data Sensitivity, Software), "
        "we use dropdowns and perform a simple inclusion/equality match with support for 'any' where available."
    )


if __name__ == "__main__":
    main()
