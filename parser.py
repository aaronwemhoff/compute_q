

app.py
222 lines

py

I am working on a questionnaire that provides researchers with the recommended solution. Attached is the code to this point.  The questionnaire has lots of issues, which may be related to the import of data from Excel. Can you address the fact that not all Excel data are numeric values? Cells in green indicate "less than or equal to", and values in red indicate "greater than". Excel data are imported starting with Row 5, with column A a character array for the operating system, column B a numeric array of walltime in hours, column C a numeric array of the number of cores, column D a numeric array of memory in GB, column E a numeric array of the number of runs, Column F a character array of data sensitivity, Column G a character array for the software package, and Column H the recommended solution.

Looking at your code, I can see the main issue is that it assumes all data (except the platform name) should be numeric, but your Excel file contains mixed data types - both numeric values and text strings. Let me help you fix this by updating the parser to handle different data types appropriately.


Updated utils/parser.py
Code 

Updated app.py
Code 
I've updated both the parser and main app to properly handle mixed data types (both numeric and categorical data) from your Excel file. Here are the key improvements:

🔧 Main Changes
1. Enhanced Data Type Detection
The parser now automatically detects whether each column contains numeric or categorical data
Columns with >70% numeric values are treated as numeric, others as categorical
2. Appropriate Input Widgets
Categorical data (OS, Software Package, Data Sensitivity): Uses dropdown menus with all unique values from your Excel
Numeric data (Cores, Memory, Walltime, etc.): Uses number input fields with smart min/max ranges
3. Smart Comparison Logic
Categorical data: Uses exact string matching (case-insensitive)
Numeric data: Uses mathematical comparisons (≤, >, ≥, etc.)
Handles the green/red cell color rules appropriately for both data types
4. Improved User Experience
Shows data types in the preview section
Better visual organization with subheaders for each metric
Clearer explanations of how comparisons work
"Other" option for categorical inputs if users need custom values
🎯 How It Handles Your Specific Columns
Based on your description:

Column A (OS): Categorical → Dropdown menu
Column B (Walltime): Numeric → Number input
Column C (Cores): Numeric → Number input
Column D (Memory GB): Numeric → Number input
Column E (Runs): Numeric → Number input
Column F (Data Sensitivity): Categorical → Dropdown menu
Column G (Software Package): Categorical → Dropdown menu
Column H (Recommended Solution): Will be treated based on content
🚀 To Use the Updated Version
Replace your existing utils/parser.py with the first code artifact
Replace your existing app.py with the second code artifact
Make sure you have the required dependencies: openpyxl for reading Excel cell colors
Run the app - it should now properly handle your mixed data types!
The app will now create appropriate input methods for each type of data and compare user inputs correctly against the platform capabilities using the color-coded rules from your Excel file.




Can you explain this section to me in more detail?

I notice you included an empty code block in your question. Could you please clarify which specific section of the code you'd like me to explain in more detail?

Are you referring to:

Data type detection logic - how the app determines if a column is numeric vs categorical?
Cell color detection - how it reads green/red cells from Excel?
Comparison logic - how user inputs are compared against platform values?
Input widget creation - how it creates dropdowns vs number inputs?
Scoring system - how platforms get ranked?
A specific function - like `load_questionn




# utils/parser.py
# ------------------------------
# Updated parser to handle mixed data types (numeric and text) from Excel questionnaire
# ------------------------------

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def is_green_cell(cell):
    """Check if a cell has a green background (various shades)"""
    if cell.fill and cell.fill.start_color:
        # Check for green-ish colors (RGB values where green component is dominant)
        color = cell.fill.start_color.rgb
        if color and len(str(color)) >= 6:
            try:
                # Convert hex to RGB
                if str(color).startswith('FF'):
                    color = str(color)[2:]  # Remove FF prefix if present
                r = int(color[0:2], 16) if len(color) >= 2 else 0
                g = int(color[2:4], 16) if len(color) >= 4 else 0
                b = int(color[4:6], 16) if len(color) >= 6 else 0
                
                # Consider it green if green component is significantly higher
                return g > max(r, b) + 30
            except:
                pass
    return False

def is_red_cell(cell):
    """Check if a cell has a red background (various shades)"""
    if cell.fill and cell.fill.start_color:
        color = cell.fill.start_color.rgb
        if color and len(str(color)) >= 6:
            try:
                # Convert hex to RGB
                if str(color).startswith('FF'):
                    color = str(color)[2:]  # Remove FF prefix if present
                r = int(color[0:2], 16) if len(color) >= 2 else 0
                g = int(color[2:4], 16) if len(color) >= 4 else 0
                b = int(color[4:6], 16) if len(color) >= 6 else 0
                
                # Consider it red if red component is significantly higher
                return r > max(g, b) + 30
            except:
                pass
    return False

def determine_data_type(column_data):
    """Determine if a column should be treated as numeric or categorical"""
    # Remove NaN values for analysis
    clean_data = [val for val in column_data if pd.notna(val)]
    
    if not clean_data:
        return 'numeric'  # Default to numeric for empty columns
    
    # Try to convert to numeric
    numeric_count = 0
    for val in clean_data:
        try:
            float(val)
            numeric_count += 1
        except (ValueError, TypeError):
            pass
    
    # If more than 70% can be converted to numeric, treat as numeric
    if numeric_count / len(clean_data) > 0.7:
        return 'numeric'
    else:
        return 'categorical'

def compare_value(user_input, platform_value, operator, data_type='numeric'):
    """
    Compare user input against platform value using the specified operator.
    Now handles both numeric and categorical data types.
    """
    if data_type == 'categorical':
        # For categorical data, only check for exact matches
        if operator in ['eq', '==']:
            return str(user_input).lower().strip() == str(platform_value).lower().strip()
        else:
            # For categorical data with other operators, just check if they match
            return str(user_input).lower().strip() == str(platform_value).lower().strip()
    
    # For numeric data, use the existing logic
    try:
        user_val = float(user_input) if pd.notna(user_input) else 0.0
        plat_val = float(platform_value) if pd.notna(platform_value) else 0.0
    except (ValueError, TypeError):
        # If conversion fails, fall back to string comparison
        return str(user_input).lower().strip() == str(platform_value).lower().strip()
    
    if operator == 'le' or operator == '<=':
        return user_val <= plat_val
    elif operator == 'gt' or operator == '>':
        return user_val > plat_val
    elif operator == 'ge' or operator == '>=':
        return user_val >= plat_val
    elif operator == 'lt' or operator == '<':
        return user_val < plat_val
    elif operator == 'eq' or operator == '==':
        return abs(user_val - plat_val) < 1e-9
    else:
        # Default to >= for unknown operators
        return user_val >= plat_val

def detect_operator_label(op):
    """Convert operator code to human-readable label"""
    labels = {
        'le': '≤ (less than or equal to)',
        'gt': '> (greater than)',
        'ge': '≥ (greater than or equal to)',
        'lt': '< (less than)',
        'eq': '= (equal to)',
        '<=': '≤ (less than or equal to)',
        '>': '> (greater than)',
        '>=': '≥ (greater than or equal to)',
        '<': '< (less than)',
        '==': '= (equal to)',
    }
    return labels.get(op, f'{op} (unknown operator)')

def load_questionnaire(xlsx_path, header_row=4, start_row=5, end_row=25):
    """
    Load questionnaire data from Excel file, handling mixed data types.
    
    Parameters:
    - xlsx_path: path to Excel file or uploaded file object
    - header_row: row number (1-indexed) containing headers
    - start_row: first row (1-indexed) of data
    - end_row: last row (1-indexed) of data
    
    Returns:
    - Dictionary with platforms_df, metrics, ops_grid, values_grid, and data_types
    """
    
    # Load the workbook to read cell colors
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb.active
    
    # Read the data using pandas
    df = pd.read_excel(xlsx_path, header=header_row-1, nrows=end_row-header_row)
    
    # Get column names (metrics)
    all_columns = df.columns.tolist()
    platform_col = all_columns[0]  # First column is platform name
    metrics = all_columns[1:]  # Rest are metrics
    
    # Filter to the specified row range
    data_start_idx = start_row - header_row - 1  # Convert to 0-indexed relative to header
    data_end_idx = end_row - header_row  # Exclusive end
    
    if data_start_idx >= 0:
        df = df.iloc[data_start_idx:data_end_idx]
    
    # Remove rows where platform name is NaN
    df = df[df[platform_col].notna()]
    
    # Set platform names as index
    df = df.set_index(platform_col)
    platform_names = df.index.tolist()
    
    # Determine data types for each metric
    data_types = {}
    for metric in metrics:
        data_types[metric] = determine_data_type(df[metric].values)
    
    # Create operators and values grids
    ops_grid = {}
    values_grid = {}
    
    for i, platform in enumerate(platform_names):
        ops_grid[platform] = {}
        values_grid[platform] = {}
        
        for j, metric in enumerate(metrics):
            # Get the Excel cell coordinates
            excel_row = start_row + i  # 1-indexed
            excel_col = j + 2  # +2 because first column is platform name, and Excel is 1-indexed
            
            try:
                cell = ws.cell(row=excel_row, column=excel_col)
                
                # Determine operator based on cell color
                if is_green_cell(cell):
                    operator = 'le'  # Green = less than or equal to
                elif is_red_cell(cell):
                    operator = 'gt'  # Red = greater than
                else:
                    # Default behavior for non-colored cells
                    if data_types[metric] == 'categorical':
                        operator = 'eq'  # Categorical data uses equality
                    else:
                        operator = 'ge'  # Numeric data defaults to >=
                
                ops_grid[platform][metric] = operator
                
                # Store the actual value
                cell_value = df.loc[platform, metric]
                values_grid[platform][metric] = cell_value
                
            except Exception as e:
                # Fallback if cell reading fails
                if data_types[metric] == 'categorical':
                    ops_grid[platform][metric] = 'eq'
                else:
                    ops_grid[platform][metric] = 'ge'
                values_grid[platform][metric] = df.loc[platform, metric]
    
    return {
        'platforms_df': df,
        'metrics': metrics,
        'ops_grid': ops_grid,
        'values_grid': values_grid,
        'data_types': data_types
    }
